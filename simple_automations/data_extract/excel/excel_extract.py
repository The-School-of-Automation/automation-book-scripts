import os

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

# mapping from index to letter
letters = [
    "A", "B", "C", "D", "E", "F", "G", "H", "I", 
    "J", "K", "L", "M", "N", "O", "P", "Q", "R", 
    "S", "T", "U", "V", "W", "X", "Y", "Z"
]

# needed information about excel workbook
summary_wb_name = "summary.xlsx"
sheet_name = "Source Data"
column = "C"
start_row = 4

# get all the files from the current folder
dir_content = os.listdir(".")
excel_tl_files = [doc for doc in dir_content if doc.endswith("xltx")]
# sorts the names by year
excel_tl_files.sort()
processed = 0

# set up the output workbook
summary_wb = Workbook()
summary_ws = summary_wb.active
summary_ws["A1"] = "Customers"
summary_ws["A1"].font = Font(size="16", bold=True)

for index, excel_file in enumerate(excel_tl_files):
    print(f"Processing and extracting data from {excel_file}")
    wb = load_workbook(filename=excel_file)
    sheet = wb[sheet_name]

    # set up the counter and get the current cell
    curr_row = start_row
    cell_num = f"{column}{curr_row}"
    cell = sheet[cell_num]

    # empty list of entities for this workbook
    workbook_entities = []

    # get distinct customers from all years
    while cell.value is not None:
        # get the value of the current cell
        entity = cell.value

        # if the entity is not yet part of the array, add it
        if entity not in workbook_entities:
            workbook_entities.append(entity)

        # set the current cell to the next row
        curr_row += 1
        cell_num = f"{column}{curr_row}"
        cell = sheet[cell_num]

    # write to new excel sheet with unique customers and heading of the given year
    column_letter = letters[index]
    column_index = f"{column_letter}2"
    summary_ws.column_dimensions[column_letter].width = 20

    # write the name of the file without filetype
    summary_ws[column_index] = os.path.splitext(excel_file)[0]
    summary_ws[column_index].font = Font(bold=True)

    # for every entity, write it to a row below the header
    for i, entity in enumerate(workbook_entities):
        cell_index = f"{column_letter}{i + 3}"
        summary_ws[cell_index] = entity

    processed += 1

# save the workbook
summary_wb.save(summary_wb_name)
print(f"Processed {processed} of {len(excel_tl_files)} excel files.")